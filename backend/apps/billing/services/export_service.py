"""
Statement export service for Excel and PDF generation.
"""

from __future__ import annotations

from collections import defaultdict
from decimal import Decimal
from io import BytesIO
from typing import TYPE_CHECKING

from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.core.services.base_service import BaseService


if TYPE_CHECKING:
    from ..models import MonthlyStatement, OnDemandInvoice, TerminalSettings
    from .storage_cost_service import ContainerCostResult


class StatementExportService(BaseService):
    """Service for exporting statements to Excel and PDF formats."""

    def export_to_excel(self, statement: "MonthlyStatement") -> BytesIO:
        """Generate Excel file from statement."""
        wb = Workbook()
        ws = wb.active
        ws.title = f"Выписка {statement.month:02d}-{statement.year}"

        # Styles
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_text = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Title section
        ws.merge_cells("A1:K1")
        ws["A1"] = f"Выписка за {statement.month_name} {statement.year}"
        ws["A1"].font = header_font
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A3"] = f"Компания: {statement.company.name}"
        ws["A4"] = f"Метод расчёта: {statement.billing_method_display}"
        ws["A5"] = f"Дата формирования: {statement.generated_at.strftime('%d.%m.%Y %H:%M')}"

        # Summary section
        ws["A7"] = "ИТОГО:"
        ws["A7"].font = subheader_font
        ws["A8"] = f"Контейнеров: {statement.total_containers}"
        ws["A9"] = f"Оплачиваемых дней: {statement.total_billable_days}"
        ws["A10"] = f"Сумма USD: ${statement.total_usd:,.2f}"
        ws["A11"] = f"Сумма UZS: {statement.total_uzs:,.0f} сум"

        # Table headers
        headers = [
            "Контейнер",
            "Размер",
            "Статус",
            "Начало",
            "Конец",
            "Всего дней",
            "Льготных",
            "Оплачиваемых",
            "Ставка USD",
            "Сумма USD",
            "Сумма UZS",
        ]

        row = 13
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_text
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Table data
        for item in statement.line_items.all():
            row += 1
            end_display = (
                "На терминале" if item.is_still_on_terminal else item.period_end.strftime("%d.%m.%Y")
            )
            data = [
                item.container_number,
                item.container_size_display,
                item.container_status_display,
                item.period_start.strftime("%d.%m.%Y"),
                end_display,
                item.total_days,
                item.free_days,
                item.billable_days,
                item.daily_rate_usd,
                item.amount_usd,
                item.amount_uzs,
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if col >= 6:  # Numbers
                    cell.alignment = Alignment(horizontal="right")
                if col in (9, 10):  # USD amounts
                    cell.number_format = "#,##0.00"
                elif col == 11:  # UZS amounts
                    cell.number_format = "#,##0"

        # Adjust column widths
        column_widths = [15, 10, 12, 12, 14, 10, 10, 12, 12, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_to_pdf(self, statement: "MonthlyStatement") -> BytesIO:
        """Generate PDF from statement using HTML template."""
        from weasyprint import HTML

        # Render HTML template
        html_content = render_to_string(
            "billing/statement_pdf.html",
            {
                "statement": statement,
                "line_items": statement.line_items.all(),
            },
        )

        # Convert to PDF
        output = BytesIO()
        HTML(string=html_content).write_pdf(output)
        output.seek(0)
        return output

    def get_excel_filename(self, statement: "MonthlyStatement") -> str:
        """Generate filename for Excel export."""
        company_slug = statement.company.slug or "company"
        return f"statement_{company_slug}_{statement.year}_{statement.month:02d}.xlsx"

    def get_pdf_filename(self, statement: "MonthlyStatement") -> str:
        """Generate filename for PDF export."""
        company_slug = statement.company.slug or "company"
        return f"statement_{company_slug}_{statement.year}_{statement.month:02d}.pdf"

    # --- On-demand invoice exports ---

    def export_on_demand_to_excel(self, invoice: "OnDemandInvoice") -> BytesIO:
        """Generate Excel file from on-demand invoice."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Разовый счёт"

        # Styles
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_text = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Title section
        ws.merge_cells("A1:K1")
        invoice_ref = invoice.invoice_number or "ЧЕРНОВИК"
        ws["A1"] = f"Разовый счёт {invoice_ref}"
        ws["A1"].font = header_font
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A3"] = f"Компания: {invoice.company.name}"
        ws["A4"] = f"Дата создания: {invoice.created_at.strftime('%d.%m.%Y %H:%M')}"
        if invoice.finalized_at:
            ws["A5"] = f"Дата утверждения: {invoice.finalized_at.strftime('%d.%m.%Y %H:%M')}"

        if invoice.notes:
            ws["A6"] = f"Примечание: {invoice.notes}"

        # Summary section
        ws["A8"] = "ИТОГО:"
        ws["A8"].font = subheader_font
        ws["A9"] = f"Контейнеров: {invoice.total_containers}"
        ws["A10"] = f"Сумма USD: ${invoice.total_usd:,.2f}"
        ws["A11"] = f"Сумма UZS: {invoice.total_uzs:,.0f} сум"

        # Storage items table
        headers = [
            "Контейнер", "Размер", "Статус", "Начало", "Конец",
            "Всего дней", "Льготных", "Оплачиваемых",
            "Ставка USD", "Сумма USD", "Сумма UZS",
        ]

        row = 13
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_text
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for item in invoice.items.all():
            row += 1
            data = [
                item.container_number,
                item.container_size,
                item.container_status,
                item.entry_date.strftime("%d.%m.%Y"),
                item.exit_date.strftime("%d.%m.%Y") if item.exit_date else "На терминале",
                item.total_days,
                item.free_days,
                item.billable_days,
                item.daily_rate_usd,
                item.amount_usd,
                item.amount_uzs,
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if col >= 6:
                    cell.alignment = Alignment(horizontal="right")
                if col in (9, 10):  # USD amounts
                    cell.number_format = "#,##0.00"
                elif col == 11:  # UZS amounts
                    cell.number_format = "#,##0"

        # Service items section (if any)
        service_items = invoice.service_items.all()
        if service_items.exists():
            row += 2
            ws.cell(row=row, column=1, value="Дополнительные начисления").font = subheader_font

            row += 1
            svc_headers = ["Контейнер", "Описание", "Дата", "Сумма USD", "Сумма UZS"]
            for col, header in enumerate(svc_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_text
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

            for svc in service_items:
                row += 1
                svc_data = [
                    svc.container_number,
                    svc.description,
                    svc.charge_date.strftime("%d.%m.%Y"),
                    svc.amount_usd,
                    svc.amount_uzs,
                ]
                for col, value in enumerate(svc_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = thin_border
                    if col >= 4:
                        cell.alignment = Alignment(horizontal="right")
                    if col == 4:  # USD
                        cell.number_format = "#,##0.00"
                    elif col == 5:  # UZS
                        cell.number_format = "#,##0"

        # Column widths
        column_widths = [15, 10, 12, 12, 14, 10, 10, 12, 12, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_on_demand_to_pdf(self, invoice: "OnDemandInvoice") -> BytesIO:
        """Generate PDF from on-demand invoice using HTML template."""
        from weasyprint import HTML

        html_content = render_to_string(
            "billing/on_demand_invoice_pdf.html",
            {
                "invoice": invoice,
                "items": invoice.items.all(),
                "service_items": invoice.service_items.all(),
            },
        )

        output = BytesIO()
        HTML(string=html_content).write_pdf(output)
        output.seek(0)
        return output

    def get_on_demand_excel_filename(self, invoice: "OnDemandInvoice") -> str:
        """Generate filename for on-demand invoice Excel export."""
        company_slug = invoice.company.slug or "company"
        ref = invoice.invoice_number or f"draft-{invoice.id}"
        return f"invoice_{company_slug}_{ref}.xlsx"

    def get_on_demand_pdf_filename(self, invoice: "OnDemandInvoice") -> str:
        """Generate filename for on-demand invoice PDF export."""
        company_slug = invoice.company.slug or "company"
        ref = invoice.invoice_number or f"draft-{invoice.id}"
        return f"invoice_{company_slug}_{ref}.pdf"

    # --- Storage costs export ---

    def export_storage_costs_to_excel(
        self,
        cost_results: list[ContainerCostResult],
        company_name: str,
    ) -> BytesIO:
        """Generate Excel file from storage cost calculation results."""
        from django.utils import timezone

        wb = Workbook()
        ws = wb.active
        ws.title = "Текущие расходы"

        # Styles
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_text = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Title
        ws.merge_cells("A1:K1")
        ws["A1"] = f"Текущие расходы — {company_name}"
        ws["A1"].font = header_font
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A3"] = f"Дата формирования: {timezone.now().strftime('%d.%m.%Y %H:%M')}"
        ws["A4"] = f"Контейнеров: {len(cost_results)}"

        # Summary
        total_usd = sum(r.total_usd for r in cost_results)
        total_uzs = sum(r.total_uzs for r in cost_results)
        total_billable = sum(r.billable_days for r in cost_results)

        ws["A6"] = "ИТОГО:"
        ws["A6"].font = subheader_font
        ws["A7"] = f"Оплачиваемых дней: {total_billable}"
        ws["A8"] = f"Сумма USD: ${total_usd:,.2f}"
        ws["A9"] = f"Сумма UZS: {total_uzs:,.0f} сум"

        # Table headers
        headers = [
            "Контейнер", "Размер", "Статус", "Дата въезда", "Дата выезда",
            "Всего дней", "Льготных", "Оплачиваемых",
            "Сумма USD", "Сумма UZS",
        ]

        row = 11
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_text
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for result in cost_results:
            row += 1
            end_display = "На терминале" if result.is_active else result.end_date.strftime("%d.%m.%Y")
            data = [
                result.container_number,
                result.container_size,
                result.container_status,
                result.entry_date.strftime("%d.%m.%Y"),
                end_display,
                result.total_days,
                result.free_days_applied,
                result.billable_days,
                result.total_usd,
                result.total_uzs,
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if col >= 6:
                    cell.alignment = Alignment(horizontal="right")
                if col == 9:  # USD
                    cell.number_format = "#,##0.00"
                elif col == 10:  # UZS
                    cell.number_format = "#,##0"

        # Column widths
        column_widths = [15, 10, 12, 12, 14, 10, 10, 12, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # --- Счёт-фактура (formal invoice) export ---

    def export_to_schet_factura(
        self,
        statement: "MonthlyStatement",
        settings: "TerminalSettings",
        exchange_rate: Decimal | None = None,
    ) -> BytesIO:
        """Generate formal Uzbek Счёт-фактура Excel from a monthly statement.

        Layout follows the standard two-part format:
          Part 1 – АКТ (acceptance act, summary)
          Part 2 – Счёт-фактура (invoice detail table)

        Each line item produces two rows: one in USD, one in UZS.
        """
        from django.utils import timezone

        wb = Workbook()
        ws = wb.active
        ws.title = "Счёт-фактура"

        rate = exchange_rate or settings.default_usd_uzs_rate or Decimal("0")

        # ----- styles -----
        bold = Font(bold=True)
        bold_center = Font(bold=True, size=11)
        small = Font(size=9)
        header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        thin = Side(style="thin")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

        def _cell(row: int, col: int, value: object, **kw: object) -> None:
            """Helper to set value + styles on a cell."""
            c = ws.cell(row=row, column=col, value=value)
            if "font" in kw:
                c.font = kw["font"]  # type: ignore[assignment]
            if "fill" in kw:
                c.fill = kw["fill"]  # type: ignore[assignment]
            if "alignment" in kw:
                c.alignment = kw["alignment"]  # type: ignore[assignment]
            if "border" in kw:
                c.border = kw["border"]  # type: ignore[assignment]
            if "number_format" in kw:
                c.number_format = kw["number_format"]  # type: ignore[assignment]

        # =====================================================================
        # PART 1 — АКТ ВЫПОЛНЕННЫХ РАБОТ (acceptance act header)
        # =====================================================================
        row = 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        _cell(row, 1, "АКТ ВЫПОЛНЕННЫХ РАБОТ", font=Font(bold=True, size=14), alignment=align_center)

        row = 3
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        period_label = f"за {statement.month_name} {statement.year} г."
        _cell(row, 1, period_label, font=bold_center, alignment=align_center)

        # Supplier / buyer block
        row = 5
        _cell(row, 1, "Исполнитель:", font=bold)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        _cell(row, 2, settings.company_name or "—")

        row = 6
        _cell(row, 1, "Заказчик:", font=bold)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        _cell(row, 2, statement.company.name)

        row = 7
        _cell(row, 1, "Основание:", font=bold)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        _cell(row, 2, "Договор хранения контейнеров")

        row = 9
        _cell(row, 1, "Курс ЦБ РУз:", font=bold)
        _cell(row, 2, f"1 USD = {rate:,.2f} сум")

        # =====================================================================
        # PART 2 — Detail table
        # =====================================================================
        # Column layout (A-J):
        #  A=№, B=Наименование, C=Ед.изм, D=Кол-во, E=Цена, F=Сумма,
        #  G=НДС, H=Всего с НДС, I=Начало, J=Конец

        table_start = 11
        th_row = table_start

        th_headers = [
            ("A", "№"),
            ("B", "Наименование услуги"),
            ("C", "Ед.изм"),
            ("D", "Кол-во"),
            ("E", "Цена"),
            ("F", "Сумма"),
            ("G", "В т.ч. НДС"),
            ("H", "Всего с НДС"),
            ("I", "Начало"),
            ("J", "Конец"),
        ]
        for col_idx, (_, header) in enumerate(th_headers, 1):
            _cell(th_row, col_idx, header, font=bold, fill=header_fill,
                  border=border_all, alignment=align_center)

        # Group line items by size+status for compact display
        grouped = self._group_line_items(statement)
        vat_rate = settings.vat_rate or Decimal("12")

        row = th_row + 1
        item_no = 0
        grand_total_usd = Decimal("0")
        grand_total_uzs = Decimal("0")
        grand_vat_usd = Decimal("0")
        grand_vat_uzs = Decimal("0")

        for label, total_usd, qty, unit, period_start, period_end in grouped:
            item_no += 1
            total_uzs = total_usd * rate

            # НДС added on top of net amount
            vat_usd = (total_usd * vat_rate / Decimal("100")).quantize(Decimal("0.01"))
            vat_uzs = (total_uzs * vat_rate / Decimal("100")).quantize(Decimal("0.01"))

            grand_total_usd += total_usd
            grand_total_uzs += total_uzs
            grand_vat_usd += vat_usd
            grand_vat_uzs += vat_uzs

            # USD row
            _cell(row, 1, item_no, border=border_all, alignment=align_center)
            _cell(row, 2, label, border=border_all, alignment=align_left)
            _cell(row, 3, unit, border=border_all, alignment=align_center)
            _cell(row, 4, qty, border=border_all, alignment=align_center)
            _cell(row, 5, float(total_usd / qty if qty else 0),
                  border=border_all, alignment=align_right, number_format="#,##0.00")
            _cell(row, 6, float(total_usd),
                  border=border_all, alignment=align_right, number_format="#,##0.00")
            _cell(row, 7, float(vat_usd),
                  border=border_all, alignment=align_right, number_format="#,##0.00")
            _cell(row, 8, float(total_usd + vat_usd),
                  border=border_all, alignment=align_right, number_format="#,##0.00")
            _cell(row, 9, period_start, border=border_all, alignment=align_center)
            _cell(row, 10, period_end, border=border_all, alignment=align_center)
            row += 1

            # UZS row (same item, converted)
            _cell(row, 1, "", border=border_all)
            _cell(row, 2, f"  в т.ч. UZS ({label})", border=border_all, alignment=align_left, font=small)
            _cell(row, 3, unit, border=border_all, alignment=align_center, font=small)
            _cell(row, 4, qty, border=border_all, alignment=align_center, font=small)
            _cell(row, 5, float(total_uzs / qty if qty else 0),
                  border=border_all, alignment=align_right, number_format="#,##0")
            _cell(row, 6, float(total_uzs),
                  border=border_all, alignment=align_right, number_format="#,##0")
            _cell(row, 7, float(vat_uzs),
                  border=border_all, alignment=align_right, number_format="#,##0")
            _cell(row, 8, float(total_uzs + vat_uzs),
                  border=border_all, alignment=align_right, number_format="#,##0")
            _cell(row, 9, "", border=border_all)
            _cell(row, 10, "", border=border_all)
            row += 1

        # Service items (if any on the statement's on-demand invoices)
        service_grouped = self._group_service_items(statement)
        for desc, svc_total_usd, svc_qty, svc_unit in service_grouped:
            item_no += 1
            svc_total_uzs = svc_total_usd * rate
            # НДС added on top of net amount
            svc_vat_usd = (svc_total_usd * vat_rate / Decimal("100")).quantize(Decimal("0.01"))
            svc_vat_uzs = (svc_total_uzs * vat_rate / Decimal("100")).quantize(Decimal("0.01"))

            grand_total_usd += svc_total_usd
            grand_total_uzs += svc_total_uzs
            grand_vat_usd += svc_vat_usd
            grand_vat_uzs += svc_vat_uzs

            _cell(row, 1, item_no, border=border_all, alignment=align_center)
            _cell(row, 2, desc, border=border_all, alignment=align_left)
            _cell(row, 3, svc_unit, border=border_all, alignment=align_center)
            _cell(row, 4, svc_qty, border=border_all, alignment=align_center)
            _cell(row, 5, float(svc_total_usd / svc_qty if svc_qty else 0),
                  border=border_all, alignment=align_right, number_format="#,##0.00")
            _cell(row, 6, float(svc_total_usd),
                  border=border_all, alignment=align_right, number_format="#,##0.00")
            _cell(row, 7, float(svc_vat_usd),
                  border=border_all, alignment=align_right, number_format="#,##0.00")
            _cell(row, 8, float(svc_total_usd + svc_vat_usd),
                  border=border_all, alignment=align_right, number_format="#,##0.00")
            _cell(row, 9, "", border=border_all)
            _cell(row, 10, "", border=border_all)
            row += 1

            # UZS row
            _cell(row, 1, "", border=border_all)
            _cell(row, 2, f"  в т.ч. UZS ({desc})", border=border_all, alignment=align_left, font=small)
            _cell(row, 3, svc_unit, border=border_all, alignment=align_center, font=small)
            _cell(row, 4, svc_qty, border=border_all, alignment=align_center, font=small)
            _cell(row, 5, float(svc_total_uzs / svc_qty if svc_qty else 0),
                  border=border_all, alignment=align_right, number_format="#,##0")
            _cell(row, 6, float(svc_total_uzs),
                  border=border_all, alignment=align_right, number_format="#,##0")
            _cell(row, 7, float(svc_vat_uzs),
                  border=border_all, alignment=align_right, number_format="#,##0")
            _cell(row, 8, float(svc_total_uzs + svc_vat_uzs),
                  border=border_all, alignment=align_right, number_format="#,##0")
            _cell(row, 9, "", border=border_all)
            _cell(row, 10, "", border=border_all)
            row += 1

        # =====================================================================
        # TOTALS ROW
        # =====================================================================
        _cell(row, 1, "", border=border_all)
        _cell(row, 2, "ИТОГО USD:", font=bold, border=border_all, alignment=align_right)
        _cell(row, 3, "", border=border_all)
        _cell(row, 4, "", border=border_all)
        _cell(row, 5, "", border=border_all)
        _cell(row, 6, float(grand_total_usd), font=bold,
              border=border_all, alignment=align_right, number_format="#,##0.00")
        _cell(row, 7, float(grand_vat_usd), font=bold,
              border=border_all, alignment=align_right, number_format="#,##0.00")
        _cell(row, 8, float(grand_total_usd + grand_vat_usd), font=bold,
              border=border_all, alignment=align_right, number_format="#,##0.00")
        _cell(row, 9, "", border=border_all)
        _cell(row, 10, "", border=border_all)
        row += 1

        _cell(row, 1, "", border=border_all)
        _cell(row, 2, "ИТОГО UZS:", font=bold, border=border_all, alignment=align_right)
        _cell(row, 3, "", border=border_all)
        _cell(row, 4, "", border=border_all)
        _cell(row, 5, "", border=border_all)
        _cell(row, 6, float(grand_total_uzs), font=bold,
              border=border_all, alignment=align_right, number_format="#,##0")
        _cell(row, 7, float(grand_vat_uzs), font=bold,
              border=border_all, alignment=align_right, number_format="#,##0")
        _cell(row, 8, float(grand_total_uzs + grand_vat_uzs), font=bold,
              border=border_all, alignment=align_right, number_format="#,##0")
        _cell(row, 9, "", border=border_all)
        _cell(row, 10, "", border=border_all)
        row += 2

        # =====================================================================
        # SIGNATURES
        # =====================================================================
        _cell(row, 1, "Исполнитель:", font=bold)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        _cell(row, 2, f"{settings.director_title or 'Руководитель'} _______________ {settings.director_name or ''}")
        _cell(row, 6, "Заказчик:", font=bold)
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=10)
        _cell(row, 7, "_________________ _______________")

        row += 2
        _cell(row, 1, "Гл. бухгалтер:", font=bold)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        _cell(row, 2, f"_______________ {settings.accountant_name or ''}")

        # Column widths
        col_widths = [5, 35, 8, 8, 14, 14, 14, 14, 12, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Save
        sf_output = BytesIO()
        wb.save(sf_output)
        sf_output.seek(0)
        return sf_output

    def get_schet_factura_filename(self, statement: "MonthlyStatement") -> str:
        """Generate filename for Счёт-фактура export."""
        company_slug = statement.company.slug or "company"
        return f"sf_{company_slug}_{statement.year}_{statement.month:02d}.xlsx"

    def export_to_schet_factura_pdf(
        self,
        statement: "MonthlyStatement",
        settings: "TerminalSettings",
        exchange_rate: Decimal | None = None,
    ) -> BytesIO:
        """Generate formal Uzbek Счёт-фактура as PDF for preview.

        Uses the same grouping logic as the Excel version but renders
        via an HTML template + WeasyPrint.
        """
        from weasyprint import HTML

        rate = exchange_rate or settings.default_usd_uzs_rate or Decimal("0")
        vat_rate = settings.vat_rate or Decimal("12")

        grouped = self._group_line_items(statement)
        service_grouped = self._group_service_items(statement)

        # Build template-friendly data
        item_no = 0
        grand_total_usd = Decimal("0")
        grand_total_uzs = Decimal("0")
        grand_vat_usd = Decimal("0")
        grand_vat_uzs = Decimal("0")

        grouped_items = []
        for label, total_usd, qty, unit, period_start, period_end in grouped:
            item_no += 1
            total_uzs = total_usd * rate
            # НДС added on top of net amount
            vat_usd = (total_usd * vat_rate / Decimal("100")).quantize(Decimal("0.01"))
            vat_uzs = (total_uzs * vat_rate / Decimal("100")).quantize(Decimal("0.01"))

            grand_total_usd += total_usd
            grand_total_uzs += total_uzs
            grand_vat_usd += vat_usd
            grand_vat_uzs += vat_uzs

            grouped_items.append({
                "number": item_no,
                "label": label,
                "unit": unit,
                "qty": qty,
                "unit_price_usd": f"{total_usd / qty if qty else 0:,.2f}",
                "total_usd": f"{total_usd:,.2f}",
                "vat_usd": f"{vat_usd:,.2f}",
                "total_with_vat_usd": f"{total_usd + vat_usd:,.2f}",
                "unit_price_uzs": f"{total_uzs / qty if qty else 0:,.0f}",
                "total_uzs": f"{total_uzs:,.0f}",
                "vat_uzs": f"{vat_uzs:,.0f}",
                "total_with_vat_uzs": f"{total_uzs + vat_uzs:,.0f}",
                "period_start": period_start,
                "period_end": period_end,
            })

        svc_items = []
        for desc, svc_total_usd, svc_qty, svc_unit in service_grouped:
            item_no += 1
            svc_total_uzs = svc_total_usd * rate
            # НДС added on top of net amount
            svc_vat_usd = (svc_total_usd * vat_rate / Decimal("100")).quantize(Decimal("0.01"))
            svc_vat_uzs = (svc_total_uzs * vat_rate / Decimal("100")).quantize(Decimal("0.01"))

            grand_total_usd += svc_total_usd
            grand_total_uzs += svc_total_uzs
            grand_vat_usd += svc_vat_usd
            grand_vat_uzs += svc_vat_uzs

            svc_items.append({
                "number": item_no,
                "label": desc,
                "unit": svc_unit,
                "qty": svc_qty,
                "unit_price_usd": f"{svc_total_usd / svc_qty if svc_qty else 0:,.2f}",
                "total_usd": f"{svc_total_usd:,.2f}",
                "vat_usd": f"{svc_vat_usd:,.2f}",
                "total_with_vat_usd": f"{svc_total_usd + svc_vat_usd:,.2f}",
                "unit_price_uzs": f"{svc_total_uzs / svc_qty if svc_qty else 0:,.0f}",
                "total_uzs": f"{svc_total_uzs:,.0f}",
                "vat_uzs": f"{svc_vat_uzs:,.0f}",
                "total_with_vat_uzs": f"{svc_total_uzs + svc_vat_uzs:,.0f}",
            })

        html_content = render_to_string(
            "billing/schet_factura_pdf.html",
            {
                "statement": statement,
                "settings": settings,
                "exchange_rate": f"{rate:,.2f}",
                "grouped_items": grouped_items,
                "service_items": svc_items,
                "grand_total_usd": f"{grand_total_usd:,.2f}",
                "grand_total_uzs": f"{grand_total_uzs:,.0f}",
                "grand_vat_usd": f"{grand_vat_usd:,.2f}",
                "grand_vat_uzs": f"{grand_vat_uzs:,.0f}",
                "grand_total_with_vat_usd": f"{grand_total_usd + grand_vat_usd:,.2f}",
                "grand_total_with_vat_uzs": f"{grand_total_uzs + grand_vat_uzs:,.0f}",
            },
        )

        output = BytesIO()
        HTML(string=html_content).write_pdf(output)
        output.seek(0)
        return output

    # ----- private helpers -----

    @staticmethod
    def _group_line_items(
        statement: "MonthlyStatement",
    ) -> list[tuple[str, Decimal, int, str, str, str]]:
        """Group statement line items by container size + status.

        Returns list of (label, total_usd, quantity, unit, period_start, period_end).
        """
        groups: dict[str, dict] = defaultdict(
            lambda: {"total_usd": Decimal("0"), "qty": 0, "start": None, "end": None}
        )
        for item in statement.line_items.all():
            key = f"Хранение контейнеров {item.container_size_display} ({item.container_status_display})"
            g = groups[key]
            g["total_usd"] += item.amount_usd
            g["qty"] += 1
            start_str = item.period_start.strftime("%d.%m.%Y")
            end_str = (
                "н.в." if item.is_still_on_terminal else item.period_end.strftime("%d.%m.%Y")
            )
            if g["start"] is None or item.period_start < g["start"]:
                g["start"] = item.period_start
                g["start_str"] = start_str
            if g["end"] is None:
                g["end_str"] = end_str
        result = []
        for label, g in groups.items():
            result.append((
                label,
                g["total_usd"],
                g["qty"],
                "конт.",
                g.get("start_str", ""),
                g.get("end_str", ""),
            ))
        return result

    @staticmethod
    def _group_service_items(
        statement: "MonthlyStatement",
    ) -> list[tuple[str, Decimal, int, str]]:
        """Group service items from on-demand invoices linked to this statement.

        Returns list of (description, total_usd, quantity, unit).
        """
        # Service items may not exist yet — safe default
        groups: dict[str, dict] = defaultdict(
            lambda: {"total_usd": Decimal("0"), "qty": 0}
        )
        try:
            for invoice in statement.company.on_demand_invoices.filter(
                status="finalized",
                created_at__year=statement.year,
                created_at__month=statement.month,
            ):
                for svc in invoice.service_items.all():
                    g = groups[svc.description]
                    g["total_usd"] += svc.amount_usd
                    g["qty"] += 1
        except Exception:
            pass  # No service items relation — that's fine
        result = []
        for desc, g in groups.items():
            result.append((desc, g["total_usd"], g["qty"], "усл."))
        return result
