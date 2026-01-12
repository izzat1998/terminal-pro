import io
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from apps.core.services import BaseService

from ..models import ContainerEntry


class ContainerEntryExportService(BaseService):
    """
    Service for exporting container entries to Excel files.
    Generates Excel with all container entry data in Russian columns.
    """

    def __init__(self):
        super().__init__()

    def export_to_excel(self, queryset) -> bytes:
        """
        Export container entries to Excel file with dual headers (English + Russian).

        Args:
            queryset: Filtered ContainerEntry queryset

        Returns:
            Excel file bytes (can be written to FileResponse)

        Raises:
            Exception: If export fails
        """
        try:
            self.logger.info(f"Starting export of {queryset.count()} container entries")

            # Optimize queryset with related data
            queryset = queryset.select_related(
                "container", "container_owner", "recorded_by"
            ).prefetch_related("crane_operations")

            # Convert queryset to list of dictionaries
            export_data = []
            for idx, entry in enumerate(queryset, start=1):
                export_data.append(self._build_row(entry, idx))

            # Define English headers (Row 1)
            english_headers = [
                "№",
                "Container number",
                "Container length",
                "Client",
                "Container Owner",
                "Cargo Name",
                "Terminal IN Date",
                "Terminal IN Modality",
                "IN Train #",
                "IN Truck / Wagon #",
                "Date of pick up",
                "Terminal OUT Modality",
                "OUT Train #",
                "OUT Truck / Wagon #",
                "Destination station",
                "Location",  # Fixed spelling
                "Date of additional crane operation",
                "Note",
                "Dwell Time (days)",
                "weight of cargo",
            ]

            # Define Russian headers (Row 2)
            russian_headers = [
                "№",
                "Номер контейнера",
                "Тип",
                "Клиент",
                "Собственник контейнера",
                "Наименование ГРУЗА",
                "Дата разгрузки на терминале",
                "транспорт при ЗАВОЗЕ",
                "Номер Поезда при ЗАВОЗЕ",
                "номер машины/ вагона при ЗАВОЗЕ",
                "Дата вывоза конт-ра с МТТ",
                "Транспорт при ВЫВОЗЕ",
                "Номер Поезда при ВЫВОЗЕ",
                "номер машины/ вагона при ВЫВОЗЕ",
                "Станция назначения",
                "Местоположение",  # Fixed spelling
                "дата дополнительной крановой операции",
                "Примечание",
                "Количество дней на хранение",
                "Тоннаж",
            ]

            # Column order for data (using Russian keys)
            column_order = [
                "№",
                "Номер контейнера",
                "Тип контейнера",
                "Клиент",
                "Собственник контейнера",
                "Наименование ГРУЗА",
                "Дата разгрузки на Базе",
                "Транспорт при ЗАВОЗЕ",
                "Номер Поезда при ЗАВОЗЕ",
                "Номер машины/вагона",
                "Дата вывоза конт-ра с МТТ",
                "Транспорт при ВЫВОЗЕ",
                "Номер Поезда при ВЫВОЗЕ",
                "Номер машины/вагона при ВЫВОЗЕ",
                "Станция назначения",
                "Местоположение",
                "Дата дополнительной крановой операции",
                "Примечание",
                "Количество дней на хранение",
                "Тоннаж",
            ]

            # Create DataFrame with proper column order
            if export_data:
                df = pd.DataFrame(export_data)
                df = df[column_order]
            else:
                # Create empty DataFrame with correct columns
                df = pd.DataFrame(columns=column_order)

            # Write to Excel in memory
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                # Write data starting from row 3 (after dual headers)
                df.to_excel(
                    writer,
                    index=False,
                    sheet_name="EMPTY_cntr_IN_OUT_STCK",
                    startrow=2,  # Start data after dual headers
                    header=False,  # We'll write headers manually
                )

                # Access the worksheet to format
                worksheet = writer.sheets["EMPTY_cntr_IN_OUT_STCK"]

                # Insert English headers in Row 1
                for col_idx, header in enumerate(english_headers, start=1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.value = header

                # Insert Russian headers in Row 2
                for col_idx, header in enumerate(russian_headers, start=1):
                    cell = worksheet.cell(row=2, column=col_idx)
                    cell.value = header

                # Define header styling colors matching the import file
                # Green for columns A-J (till "IN Truck / Wagon #")
                green_fill = PatternFill(
                    start_color="92D050",  # Light green
                    end_color="92D050",
                    fill_type="solid",
                )
                # Yellow for columns K-R (from "Date of pick up" till "Note")
                yellow_fill = PatternFill(
                    start_color="FFFF00",  # Yellow
                    end_color="FFFF00",
                    fill_type="solid",
                )
                # Orange for column S ("Dwell Time (days)")
                orange_fill = PatternFill(
                    start_color="FFC000",  # Orange
                    end_color="FFC000",
                    fill_type="solid",
                )
                # Black for column T ("weight of cargo")
                black_fill = PatternFill(
                    start_color="000000",  # Black
                    end_color="000000",
                    fill_type="solid",
                )

                dark_font = Font(
                    bold=True, color="000000"
                )  # Dark text for light backgrounds
                white_font = Font(
                    bold=True, color="FFFFFF"
                )  # White text for dark backgrounds
                header_alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

                # Define column ranges
                green_range = range(1, 11)  # Columns A-J (1-10)
                yellow_range = range(11, 19)  # Columns K-R (11-18)
                orange_col = 19  # Column S
                black_col = 20  # Column T

                # Apply styling to English headers (Row 1)
                for col_idx in range(1, len(english_headers) + 1):
                    cell = worksheet.cell(row=1, column=col_idx)

                    if col_idx in green_range:
                        cell.fill = green_fill
                        cell.font = dark_font
                    elif col_idx in yellow_range:
                        cell.fill = yellow_fill
                        cell.font = dark_font
                    elif col_idx == orange_col:
                        cell.fill = orange_fill
                        cell.font = dark_font
                    elif col_idx == black_col:
                        cell.fill = black_fill
                        cell.font = white_font

                    cell.alignment = header_alignment

                # Apply styling to Russian headers (Row 2)
                for col_idx in range(1, len(russian_headers) + 1):
                    cell = worksheet.cell(row=2, column=col_idx)

                    if col_idx in green_range:
                        cell.fill = green_fill
                        cell.font = dark_font
                    elif col_idx in yellow_range:
                        cell.fill = yellow_fill
                        cell.font = dark_font
                    elif col_idx == orange_col:
                        cell.fill = orange_fill
                        cell.font = dark_font
                    elif col_idx == black_col:
                        cell.fill = black_fill
                        cell.font = white_font

                    cell.alignment = header_alignment

                # Set row heights for better readability
                worksheet.row_dimensions[1].height = 25  # English header row
                worksheet.row_dimensions[
                    2
                ].height = 30  # Russian header row (may need more space)

                # Freeze both header rows
                worksheet.freeze_panes = "A3"

                # Auto-adjust column widths with better spacing
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except (TypeError, AttributeError):
                            pass
                    # Add extra padding for readability
                    adjusted_width = min(max_length + 3, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Add borders to all cells (headers and data)
                thin_border = Border(
                    left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000"),
                )

                # Apply borders to header rows and all data rows
                for row in worksheet.iter_rows(
                    min_row=1,
                    max_row=worksheet.max_row,
                    min_col=1,
                    max_col=len(english_headers),
                ):
                    for cell in row:
                        cell.border = thin_border

            buffer.seek(0)
            file_bytes = buffer.getvalue()

            self.logger.info(
                f"Successfully exported {len(export_data)} container entries to Excel"
            )
            return file_bytes

        except Exception as e:
            self.logger.error(f"Error exporting container entries to Excel: {e!s}")
            raise

    def _build_row(self, entry: ContainerEntry, row_num: int) -> dict[str, Any]:
        """
        Build a single row of export data from a ContainerEntry.

        Args:
            entry: ContainerEntry instance
            row_num: Row number for display

        Returns:
            Dictionary with all export columns
        """
        # Format dates
        entry_time_str = (
            entry.entry_time.strftime("%Y-%m-%d %H:%M:%S") if entry.entry_time else ""
        )
        exit_date_str = (
            entry.exit_date.strftime("%Y-%m-%d %H:%M:%S") if entry.exit_date else ""
        )

        # Format crane operations - concatenate with "; "
        crane_op_dates = []
        if entry.crane_operations.exists():
            crane_op_dates = [
                op.operation_date.strftime("%Y-%m-%d %H:%M:%S")
                for op in entry.crane_operations.all()
            ]
        crane_operations_str = "; ".join(crane_op_dates) if crane_op_dates else ""

        # Get container owner name (handle null)
        container_owner_name = (
            entry.container_owner.name if entry.container_owner else ""
        )

        # Get transport type display names (Russian)
        transport_type_display = (
            entry.get_transport_type_display() if entry.transport_type else ""
        )
        exit_transport_type_display = (
            entry.get_exit_transport_type_display() if entry.exit_transport_type else ""
        )

        # Get cargo weight (convert Decimal to string, handle null)
        cargo_weight = ""
        if entry.cargo_weight:
            try:
                # Convert Decimal to float for Excel
                cargo_weight = float(entry.cargo_weight)
            except (ValueError, TypeError):
                cargo_weight = ""

        # Prefer company name over client_name
        client_display = ""
        if entry.company:
            client_display = entry.company.name
        elif entry.client_name:
            client_display = entry.client_name

        # Build row dictionary
        return {
            "№": row_num,
            "Номер контейнера": entry.container.container_number or "",
            "Тип контейнера": entry.container.iso_type or "",
            "Клиент": client_display,
            "Собственник контейнера": container_owner_name,
            "Наименование ГРУЗА": entry.cargo_name or "",
            "Дата разгрузки на Базе": entry_time_str,
            "Транспорт при ЗАВОЗЕ": transport_type_display,
            "Номер Поезда при ЗАВОЗЕ": entry.entry_train_number or "",
            "Номер машины/вагона": entry.transport_number or "",
            "Дата вывоза конт-ра с МТТ": exit_date_str,
            "Транспорт при ВЫВОЗЕ": exit_transport_type_display,
            "Номер Поезда при ВЫВОЗЕ": entry.exit_train_number or "",
            "Номер машины/вагона при ВЫВОЗЕ": entry.exit_transport_number or "",
            "Станция назначения": entry.destination_station or "",
            "Местоположение": entry.location or "",  # Fixed spelling
            "Дата дополнительной крановой операции": crane_operations_str,
            "Примечание": entry.note or "",
            "Количество дней на хранение": entry.dwell_time_days
            if entry.dwell_time_days
            else "",
            "Тоннаж": cargo_weight,
        }
